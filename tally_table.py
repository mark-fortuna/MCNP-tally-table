# TO DO:
#	- Give option to select multiple files, and to read all of them. Se link bellow:
# 	https://stackoverflow.com/questions/16790328/open-multiple-filenames-in-tkinter-and-add-the-filesnames-to-a-list
#   - The tallies have to be in order now. For example: if you enter gamma
#       tallies before neutron tallies, the program can't  read the tallies.
#       The benefit now is that the script doesn't read the whole MCNP output.
#       A solution may be to first read th MCNP output to the first tally and
#       then read from the first tally every time.
#	- Select from different default_tallies. Make a drop-down menu,
# 		where you can choose from different files with default tallies.
# 		All would have the prefix 'default_tallies' and then a name.
# 		For example: default_tallies-flagged, -all, -gamma, -neutron.
#			- Add the option to read all cell tallies.
#	- Make script more universal. - Compatible with time tallies,
#		surface tallies and other types of tallies.
#		Read the MCNP manual, to see all the different types of tallies.
#		The tally type would be selected in dropdown menu, or pre-defined
#		in first line of default_tallies file.
# 	- All frames inside tab should have the same pading. Widths are different
#		on macOS compared to Windows.
# 		The tabs themselves should alse have a uniform padding.
# 		These pad_sizes should be defined in script paramters section.
#	- The open file location button doesn't work on MacOS


import tkinter as tk	# for GUI
import tkinter.filedialog as fd		# for choosing files
import re
from turtle import left				# for finding tallies
import matplotlib.pyplot as plt
import subprocess		# for opening file locaitons in Windows file eplorer
import os				# for getting current location
import tkinter.scrolledtext as st
from numpy import right_shift	# for scrollable tally table
import openpyxl as xl	# for exporting to spreadsheet
from openpyxl.styles import Border, Font, Alignment, Side		# for a nicer spreadsheet
from string import ascii_uppercase	# for navigating spreadsheet

# SCRIPT PARAMETERS
# **************************************************************************
pad_in = 10
pad_out = 4
# pad_frame = 20
default_tallies_separator = '\n' # each tally in new line
tally_entry_width = 30
tab_width = 310
flagged_tally_specifier = '*'
spreadsheet_borders = 'thin'
# **************************************************************************

path_to_files = os.getcwd()
root = tk.Tk()
root.title('Tally to table')

debug = False	# debugging is off until changed in GUI
list_of_tabs = []

class Tab():
	def __init__(self, tab_number):
		# Variables
		self.number = tab_number
		self.state = True	   # False when tab is closed
		# Tally variables
		self.lines = []		 # lines of MCNP input file
		self.input_name = ''	# name of input file
		self.tally = []		 # tally numbers, counts and errors
		self.label = ''		 # file name w/o directory
		### Tab frame ###
		self.tab = tk.LabelFrame(root)
		self.tab.grid(row = 0, column = 2*self.number,
			columnspan=2, sticky='n')
		# Label
		self.tab_label = tk.Label(self.tab, text=f'Tab {self.number}')
		self.tab_label.grid(row=0, column=0, sticky='w')
		# Close button
		self.close_button = tk.Button(self.tab, text='\u2717',
			bg='red', command=self.Close)
		self.close_button.grid(row = 0, column=1, sticky='e')
		#
		## Floating frame ##
		self.floating_frame = tk.Frame(root)
		self.floating_frame.grid(row=0, column=2*self.number+2, sticky='n')
		# vertical space
		tk.Label(self.floating_frame, text=f'{self.number}', width=10, heigh=2).grid(row=0,column=0)
		# New tab button
		self.new_tab_button = tk.Button(self.floating_frame, text='New tab',
			bg='green', command=make_new_tab)
		self.new_tab_button.grid(row=1,column=0,sticky='n',padx=pad_out,pady=pad_out)
		# vertical space
		tk.Label(self.floating_frame, text='', width=10, heigh=8).grid(row=2,column=0)
		# Plot button
		self.plot_button = tk.Button(self.floating_frame, text='Plot', padx=pad_in, pady=pad_in, command=plot_all)
		self.plot_button.grid(row=3,column=0,padx=pad_out,pady=pad_out)
		# Make spreadsheet button
		self.make_spreadsheet = tk.Button(self.floating_frame, text='Make spreadsheet',
			padx=pad_in, pady=pad_in, command=make_spreadsheet)
		self.make_spreadsheet.grid(row=4,column=0,padx=pad_out,pady=pad_out)
		# vertical space
		tk.Label(self.floating_frame, text='', width=10, heigh=9).grid(row=5,column=0)
		# Open file locations button
		self.open_file_location_button = tk.Button(self.floating_frame,
			text='Open file location', padx=pad_in,pady=pad_in, command=open_file_locations)
		self.open_file_location_button.grid(row=6,column=0,padx=pad_out,pady=pad_out)
		# Debug button
		global debug
		if debug: d_text = 'Debug  ON'
		else: d_text = 'Debug OFF'
		self.debug_button = tk.Button(self.floating_frame, text=d_text, padx=pad_in,pady=pad_in, command=debug_toggle)
		self.debug_button.grid(row=7,column=0,padx=pad_out,pady=pad_out)
		#
		## File frame ##
		self.file_frame = tk.LabelFrame(self.tab, text='',width=tab_width,height=70)
		self.file_frame.grid(row=2, column=0, columnspan=2)
		self.file_frame.pack_propagate(0)
		# Select file button
		self.select_button = tk.Button(self.file_frame, text='Select MCNP output file',
			padx=pad_in,pady=pad_in, command=self.OpenFile)
		self.select_button.pack(padx=pad_out,pady=pad_out)
		#
		## Entry frame ##
		self.entry_frame = tk.LabelFrame(self.tab, text='', padx=2, pady=2)
		self.entry_frame.grid(row=3, column=0, columnspan=2, padx=5, pady=5)
		tk.Label(self.entry_frame,
			text='Enter tallies of interest:').grid(row=0,column=0, sticky='w')
		self.def_tally = tk.Entry(self.entry_frame, width = tally_entry_width)
		self.def_tally.grid(row=1, column=0, padx=5, pady=5)
		self.def_tally.insert(0, default_tallies)
		tk.Label(self.entry_frame,
			text="Default tallies are read from 'default_tallies' file.").grid(row=2,
				column=0)
		# Read button
		self.read_button = tk.Button(self.tab, text='Read tallies',
			padx=pad_in,pady=pad_in, command=self.ReadTally)
		self.read_button.grid(row=4, column=0, columnspan=2, padx=pad_out,pady=pad_out)
		#
		## Table frame ##
		self.table_frame = tk.LabelFrame(self.tab, text='TALLY TABLE',
			width=tab_width, height=300, padx=10,pady=10)
		# self.table_frame = tk.Canvas(self.tab,width=tab_width, height=300)
		self.table_frame.grid(row=5, column=0, columnspan=2)
		self.table_frame.pack_propagate(0)
		# content of Table frame is scroled text
		self.text_area = st.ScrolledText(self.table_frame,
                            width = 50, 
                            height = 20, 
                            font = ("Times New Roman",
                                    12))
		self.text_area.pack() # grid(column = 0, pady = 10, padx = 10)

		if debug: print(f' (+) Created tab number: {self.number}')
	# METHODS
	def GetTabNumber(self):
		return self.number
	def Close(self):
		self.close_button.destroy()
		self.tab.destroy()
		# only destroy new tab button if self.tab isnt the furthest right tab
		# print(f'Is tab {self.number} latest tab?\t{latest_tab(self)}')
		if not latest_tab(self):
			self.floating_frame.destroy() # destroy old button and plot button
		self.tab_label.destroy()
		if debug: print(f' (\u2717) Closed tab number: {self.number}')
		self.state = False
	def DeleteFloatingFrame(self):
		self.floating_frame.destroy()
	# tally specific methods
	def OpenFile(self):
		self.input_name = fd.askopenfilename(defaultextension='.o',
			title='Select MCNP output file')
		self.label = self.input_name.split('/')[-1]
		f = open(self.input_name, 'r')
		self.lines = f.readlines()
		f.close()
		if debug: print(f'Opened file on tab {self.number}.\nIt has {len(self.lines)} lines.')
		# rename tab label
		self.tab_label.destroy()
		self.tab_label = tk.Label(self.tab,
			text=f'Tab {self.number}, File: {self.label}')
		self.tab_label.grid(row=0, column=0, sticky='w')
		self.select_button.destroy()
		tk.Label(self.file_frame, text='MCNP output file:').pack(anchor='w')
		tk.Label(self.file_frame, text=self.input_name, wraplength=tab_width).pack()
	def ReadTally(self):
		input_tally = self.def_tally.get()
		# we must convert input_tally to string list of tally numbers!
		tally_number = input_tally.split(' ')
		N = len(tally_number)
		# matrix with tally name, tally count and tally error
		self.tally = [[i for i in tally_number],[None for i in tally_number],[None for i in tally_number]]
		# find and extract tally
		i = 0
		for k in range(len(tally_number)):
			# if tally number begins with flagged_tally_specifier
			# we need to look for it differently - see (A)
			tally_number_prefix = tally_number[k].split(flagged_tally_specifier)
			flagged = False
			if len(tally_number_prefix) > 1:
				flagged = True
				tally_number[k] = tally_number_prefix[1]
			form = '1tally(\s+)' + tally_number[k] + '(.*)'
			found = None
			while found == None:
				found = re.search(form, self.lines[i])
				i += 1
			# now we know at wich line tally info begins
			# the line before tally count begins with ' cell '
			#	(A) unless it is a flagged tally
			#		then it begins with 'flagged tallies'
			# the line with tally count begins with whitespace
			# tally count and error are sepperated with white space
			if flagged: form2 = 'flagged tallies'
			else: form2 = '\scell(\s+)'
			found2 = None
			while found2 == None:
				found2 = re.search(form2, self.lines[i])
				i += 1
			# line i has tally count and tally error
			t = re.split('\s+',self.lines[i])
			self.tally[1][k] = float(t[1]) # tally count
			self.tally[2][k] = float(t[2]) # tally error
		if debug: print(f'Read tallies on tab {self.number}.')
		# display tally
		_string = 'Tally\tCount\t\tError'
		for i in range(N):
			# _string += f'\n{self.tally[0][i]}\t{self.tally[1][i]}\t{self.tally[2][i]}'
			_string += '\n{}\t{:.3E}\t\t{}'.format( self.tally[0][i],
				self.tally[1][i], self.tally[2][i] )
		self.text_area.insert(tk.INSERT, _string)
		# destroy read tally buton
		self.read_button.destroy()
		if debug: print(f'Displayed tally table on tab {self.number}.')
		# Write to .txt button
		self.write_button = tk.Button(self.tab, text='Write to .txt file',
			padx=pad_in,pady=pad_in, command=self.WriteToFile)
		self.write_button.grid(row=6, column=0, padx=pad_out,pady=pad_out, sticky='e')
		# Save as ... button
		self.save_as_button = tk.Button(self.tab, text='Save as ...',
			padx=pad_in,pady=pad_in, command=self.SaveAs)
		self.save_as_button.grid(row=6, column=1, padx=pad_out,pady=pad_out, sticky='w')
	def WriteToFile(self):
		name = self.input_name.split('.')[0]
		output = name + '-tally_table.txt'
		o = open(output, 'w')
		lines = [None for i in self.tally[0]]
		for i in range(len(lines)):
			lines[i] = self.tally[0][i] + '\t' + '{:e}'.format(self.tally[1][i]) + '\t' + '{}'.format(self.tally[2][i]) + '\n'
		o.writelines(lines)
		o.close()
		if debug: print(f'Written tally table (from tab {self.number}) to .txt file.')
	def SaveAs(self):
		output_name = fd.asksaveasfilename(defaultextension='.txt', title='Save as ...')
		o2 = open(output_name, 'w')
		lines = [None for i in self.tally[0]]
		for i in range(len(lines)):
			lines[i] = self.tally[0][i] + '\t' + '{:e}'.format(self.tally[1][i]) + '\t' + '{}'.format(self.tally[2][i]) + '\n'
		o2.writelines(lines)
		o2.close()
		if debug: print(f'Saved tally table (from tab {self.number}) to file {output_name}')
	def Plot(self):
		# print(len(self.tally[0]))
		if self.tally == []:
			print(f'No tallies to plot on tab {self.number}.')
		else:
			x = [self.tally[0][i] for i in range(len(self.tally[0]))]
			y = [self.tally[1][i] for i in range(len(self.tally[0]))]
			re = [self.tally[2][i] for i in range(len(self.tally[0]))]  # relative y error
			ye = [i*j for i,j in zip(y, re)]	# absolute y error
			if debug:
				print(f'Data to plot (from tab {self.number}):')
				print(f'x:\t{x}\ny:\t{y}\nye:\t{ye}')
			self.plot = plt.errorbar(x,y,yerr=ye,label=f'{self.label}',
				capsize=2.0, ls='')
			plt.xticks(rotation=90)
			plt.tight_layout()
	def IsOpen(self):
		return self.state
	def DebugButtonToggle(self):
		# changes label on debug button
		global debug
		if debug: d_text = 'Debug  ON'
		else: d_text = 'Debug OFF'
		self.debug_button = tk.Button(self.floating_frame, text=d_text, padx=pad_in,pady=pad_in, command=debug_toggle)
		self.debug_button.grid(row=7,column=0,padx=pad_out,pady=pad_out)


# this function is not in class Tab() because it
# uses a global variable - a list of all tabs
def make_new_tab():
	global list_of_tabs
	numbers = []
	if debug: print(f'Called function make_new_tab().')
	if list_of_tabs == []:
		list_of_tabs.append(Tab(0))
	else:
		for i,v in enumerate(list_of_tabs):
			numbers.append(v.GetTabNumber())
			v.DeleteFloatingFrame()
		largest_number = max(numbers)
		list_of_tabs.append(Tab(largest_number+1))

# check if tab is the latest tab:
def latest_tab(tab):
	global list_of_tabs
	numbers = []
	for i,v in enumerate(list_of_tabs):
		numbers.append(v.GetTabNumber())
	if tab.GetTabNumber() == max(numbers): return True
	else: return False

def plot_all():
	global list_of_tabs
	if debug: print(f'Will plot data from open tabs.')
	for i,v in enumerate(list_of_tabs):
		if v.IsOpen():
			if debug: print(f'Will plot data from tab number {v.GetTabNumber()}.')
			v.Plot()
			if debug: print(f'Plotted from tab number {v.GetTabNumber()}.')
	plt.yscale('log')
	plt.xlabel('tallies')
	plt.legend()
	if debug: print(f'Will show figure.')
	plt.show()
	if debug: print(f'Have shown figure.')

def debug_toggle():
	print('Debug button has been pressed.')
	global debug
	print(f'Debug before switch: {debug}')
	if debug: debug = False
	else: debug = True
	print(f'Debug after switch: {debug}')
	# To toggle text on debug button you have to change it inside floating frame
	# Floating frame is asigned to latest tab object
	global list_of_tabs
	list_of_tabs[-1].DebugButtonToggle()

def open_file_locations():
	subprocess.Popen(f'explorer "{path_to_files}"')

def make_spreadsheet():
	# this function will be called when the make spreadsheet button is pressed
	if debug: print('Make spreadsheet button has been pressed.')
	work_book = xl.Workbook()
	if debug: print('Made spreadsheet work book.')
	work_sheet = work_book.active
	# run loop over list of open tabs
	for i,v in enumerate(list_of_tabs):
		if v.IsOpen(): #if tab is open save data to spreadsheet
			# first lets make the title and column names
			start_column = ascii_uppercase[3*i]
			end_column = ascii_uppercase[3*i+2]
			work_sheet.merge_cells(f'{start_column}1:{end_column}1')
			merged_cell = work_sheet.cell(row=1,column=3*i+1)
			merged_cell.value = v.label
			work_sheet.cell(row=2,column=3*i+1).value = 'Tally'
			work_sheet.cell(row=2,column=3*i+2).value = 'Count'
			work_sheet.cell(row=2,column=3*i+3).value = 'Error'
			# then input the data values
			for k in range(len(v.tally[0])):
				tn = v.tally[0][k]			# tally number
				tc = v.tally[1][k]			# tally count
				te = v.tally[2][k]			# tally error
				work_sheet.cell(row=k+3,column=3*i+1).value = tn
				work_sheet.cell(row=k+3,column=3*i+2).value = tc
				work_sheet.cell(row=k+3,column=3*i+2).number_format = '0.00E+00'
				work_sheet.cell(row=k+3,column=3*i+3).value = te
			# make vertical borders
			left_column = work_sheet.column_dimensions[start_column]
			right_column = work_sheet.column_dimensions[end_column]
			left_column.border = Border(left = Side(border_style = spreadsheet_borders,
				color="00000000"))
			right_column.border = Border(right = Side(border_style = spreadsheet_borders,
				color="00000000"))
	# Make first bold and centered
	first_row = work_sheet.row_dimensions[1]
	first_row.font = Font(name = 'Calibri', bold = True)
	first_row.alignment = Alignment(horizontal = 'center', vertical = 'center')
	# Make line bellow 2nd row
	second_row = work_sheet.row_dimensions[2]
	second_row.border = Border(bottom = Side(border_style = spreadsheet_borders,
		color="00000000"))
	if debug: print('Data hes been writen to spreadsheet.')
	# Save spreadsheet
	spreadsheet_name = fd.asksaveasfilename(defaultextension='.xlsx', title='Save spreadsheet as ...')
	work_book.save(spreadsheet_name)
	if debug: print('Spreadsheet has been saved.')

# see if 'default_tallies' exists
# if it doesn't leave entry field blank
try:
	with open('default_tallies', 'r') as f:
		print('default_tallies file present')
		dt = open('default_tallies', 'r')
		default_tallies = dt.read().split(default_tallies_separator)
except FileNotFoundError:
	print('default_tallies file is not present')
	default_tallies = ''


make_new_tab()


root.mainloop()
